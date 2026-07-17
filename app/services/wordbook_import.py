from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

from sqlmodel import Session, select

from app import crud, models
from app.services import dictionary

_WORD_SPLIT = re.compile(r"[\t,;|]+")
_MAX_BULK_ENRICH = 120


def entry_from_lookup(word: str, session: Session, *, enrich: bool = False) -> dict:
    cleaned = (word or "").strip()
    if not cleaned:
        return {}
    payload = (
        dictionary.lookup_word_enrich(cleaned, session=session)
        if enrich
        else dictionary.lookup_word_fast(cleaned, session=session)
    )
    translation = (payload.get("translation") or payload.get("youdao_translation") or "").strip()
    return {
        "word": cleaned,
        "lemma": payload.get("lemma") or cleaned,
        "definition": (payload.get("definition") or "").strip(),
        "translation": translation or None,
        "pronunciation": (payload.get("pronunciation") or "").strip() or None,
        "part_of_speech": (payload.get("part_of_speech") or "").strip() or None,
        "example": (payload.get("example") or "").strip() or None,
        "tags": [],
        "level": None,
    }


def _normalize_row(raw: dict) -> dict | None:
    word = str(raw.get("word") or raw.get("Word") or raw.get("单词") or "").strip()
    if not word:
        return None
    translation = str(
        raw.get("translation")
        or raw.get("Translation")
        or raw.get("中文")
        or raw.get("释义")
        or ""
    ).strip()
    pronunciation = str(
        raw.get("pronunciation")
        or raw.get("phonetic")
        or raw.get("音标")
        or ""
    ).strip()
    definition = str(raw.get("definition") or raw.get("Definition") or raw.get("英文") or "").strip()
    return {
        "word": word,
        "lemma": str(raw.get("lemma") or word).strip() or word,
        "definition": definition,
        "translation": translation or None,
        "pronunciation": pronunciation or None,
        "part_of_speech": str(raw.get("part_of_speech") or raw.get("pos") or "").strip() or None,
        "example": str(raw.get("example") or raw.get("例句") or "").strip() or None,
        "tags": raw.get("tags") or [],
        "level": raw.get("level"),
    }


def _parse_line(line: str) -> dict | None:
    text = line.strip()
    if not text or text.startswith("#"):
        return None
    if "\t" in text:
        parts = text.split("\t")
    elif "," in text and not text.count(" ") > 2:
        parts = next(csv.reader([text]))
    else:
        parts = _WORD_SPLIT.split(text)
    word = (parts[0] if parts else "").strip()
    if not word:
        return None
    translation = (parts[1] if len(parts) > 1 else "").strip()
    pronunciation = (parts[2] if len(parts) > 2 else "").strip()
    return {
        "word": word,
        "lemma": word,
        "definition": "",
        "translation": translation or None,
        "pronunciation": pronunciation or None,
        "part_of_speech": None,
        "example": None,
        "tags": [],
        "level": None,
    }


def parse_wordbook_text(content: str, filename: str = "") -> list[dict]:
    suffix = Path(filename or "").suffix.lower()
    text = content.strip()
    if not text:
        return []

    if suffix == ".json":
        payload = json.loads(text)
        if isinstance(payload, dict):
            payload = payload.get("entries") or payload.get("words") or payload.get("items") or []
        if not isinstance(payload, list):
            raise ValueError("JSON 文件需为词条数组，或包含 entries/words 字段")
        rows = [_normalize_row(item) for item in payload if isinstance(item, dict)]
        return [row for row in rows if row]

    if suffix == ".csv":
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames:
            rows = [_normalize_row(row) for row in reader]
            parsed = [row for row in rows if row]
            if parsed:
                return parsed
        return [row for line in text.splitlines() if (row := _parse_line(line))]

    lines = text.splitlines()
    return [row for line in lines if (row := _parse_line(line))]


def parse_wordbook_bytes(data: bytes, filename: str) -> list[dict]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("缺少 openpyxl 依赖，无法解析 Excel 文件") from exc
        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        headers = [str(cell or "").strip().lower() for cell in header]
        header_words = {"word", "单词", "english", "英文", "translation", "中文", "释义", "pronunciation", "phonetic", "音标"}
        has_header = any(h in header_words for h in headers if h)
        parsed: list[dict] = []
        if has_header:
            word_idx = next((i for i, h in enumerate(headers) if h in {"word", "单词", "english", "英文"}), 0)
            trans_idx = next((i for i, h in enumerate(headers) if h in {"translation", "中文", "释义", "meaning"}), -1)
            phon_idx = next((i for i, h in enumerate(headers) if h in {"pronunciation", "phonetic", "音标"}), -1)
            data_rows = rows_iter
        else:
            word_idx, trans_idx, phon_idx = 0, 1, 2
            data_rows = iter([header, *rows_iter])
        for row in data_rows:
            if not row:
                continue
            word = str(row[word_idx] or "").strip() if word_idx < len(row) else ""
            if not word:
                continue
            translation = str(row[trans_idx] or "").strip() if trans_idx >= 0 and trans_idx < len(row) else ""
            pronunciation = str(row[phon_idx] or "").strip() if phon_idx >= 0 and phon_idx < len(row) else ""
            parsed.append(
                {
                    "word": word,
                    "lemma": word,
                    "definition": "",
                    "translation": translation or None,
                    "pronunciation": pronunciation or None,
                    "part_of_speech": None,
                    "example": None,
                    "tags": [],
                    "level": None,
                }
            )
        return parsed

    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            text = ""
            continue
    else:
        raise ValueError("无法识别文件编码")
    return parse_wordbook_text(text, filename)


def enrich_entries(entries: list[dict], session: Session, *, enrich: bool = True) -> list[dict]:
    enriched: list[dict] = []
    enrich_budget = _MAX_BULK_ENRICH
    for entry in entries:
        word = (entry.get("word") or "").strip()
        if not word:
            continue
        has_translation = bool((entry.get("translation") or "").strip())
        has_pronunciation = bool((entry.get("pronunciation") or "").strip())
        if has_translation and has_pronunciation:
            enriched.append(entry)
            continue
        should_enrich = enrich and enrich_budget > 0 and not has_translation
        lookup = entry_from_lookup(word, session, enrich=should_enrich)
        if should_enrich:
            enrich_budget -= 1
        merged = dict(entry)
        for key in ("definition", "translation", "pronunciation", "part_of_speech", "example", "lemma"):
            if not (merged.get(key) or "").strip() if isinstance(merged.get(key), str) else merged.get(key):
                if lookup.get(key):
                    merged[key] = lookup[key]
        enriched.append(merged)
    return enriched


def add_word_to_wordbook(session: Session, wordbook_id: int, word: str) -> models.WordBookEntry:
    cleaned = (word or "").strip()
    if not cleaned:
        raise ValueError("单词不能为空")
    entry_data = entry_from_lookup(cleaned, session, enrich=True)
    crud.add_wordbook_entries(session, wordbook_id, [entry_data])
    row = session.exec(
        select(models.WordBookEntry).where(
            models.WordBookEntry.wordbook_id == wordbook_id,
            models.WordBookEntry.word == crud._normalize_word(cleaned),
        )
    ).first()
    if not row:
        raise RuntimeError("词条写入失败")
    return row


def import_wordbook_file(
    session: Session,
    wordbook_id: int,
    data: bytes,
    filename: str,
    *,
    enrich: bool = True,
) -> dict:
    entries = parse_wordbook_bytes(data, filename)
    if not entries:
        raise ValueError("文件中没有可导入的单词")
    entries = enrich_entries(entries, session, enrich=enrich)
    count = crud.add_wordbook_entries(session, wordbook_id, entries)
    return {
        "ok": True,
        "count": count,
        "parsed": len(entries),
        "filename": filename,
    }
